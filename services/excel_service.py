"""
excel_service.py
Geração de relatório Excel (.xlsx) do histórico de análises de crédito.

Estrutura do arquivo:
  - Aba "Histórico"     : uma linha por análise com colunas formatadas
  - Aba "Resumo"        : KPIs do período (total, aprovados, reprovados,
                          taxa de aprovação, VGV, ticket médio) por analista

Depende apenas de openpyxl (sem pandas) para manter o ambiente leve.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# ── Cores Paulo Bio ──────────────────────────────────────────────────────────
_ORANGE   = "F47920"
_NAVY     = "2C3E50"
_GREEN    = "27AE60"
_RED      = "E74C3C"
_YELLOW   = "F1C40F"
_LIGHT_BG = "F8F9FA"
_WHITE    = "FFFFFF"
_GRAY     = "BDC3C7"

# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=_NAVY, size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _border() -> Border:
    thin = Side(style="thin", color=_GRAY)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _header_style(ws, row: int, col: int, value: str, width: float | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(_NAVY)
    cell.font = _font(bold=True, color=_WHITE, size=10)
    cell.alignment = _center()
    cell.border = _border()
    if width and ws.column_dimensions[get_column_letter(col)].width < width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _data_cell(ws, row: int, col: int, value: Any, align="left", color: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(color=color or _NAVY)
    cell.alignment = _center() if align == "center" else _left()
    cell.border = _border()
    # Zebra
    if row % 2 == 0:
        cell.fill = _fill(_LIGHT_BG)


# ── Formatação de valores ─────────────────────────────────────────────────────

def _fmt_brl(valor: float | None) -> str:
    if valor is None:
        return "—"
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_data(iso: str | None) -> str:
    if not iso:
        return "—"
    try:
        return datetime.fromisoformat(iso.replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(iso)


def _cor_status(status: str) -> str:
    s = str(status).upper()
    if "REPROVADO" in s:
        return _RED
    if "RESSALVA" in s:
        return _YELLOW
    if "APROVADO" in s:
        return _GREEN
    return _NAVY


# ── Aba Histórico ─────────────────────────────────────────────────────────────

_COLUNAS_HISTORICO = [
    ("Data",        18),
    ("Empresa",     30),
    ("CNPJ",        18),
    ("Imóvel",      24),
    ("Aluguel",     14),
    ("Status",      22),
    ("Analista",    20),
]


def _construir_aba_historico(ws, registros: list[dict]) -> None:
    ws.title = "Histórico"
    ws.freeze_panes = "A2"  # congela cabeçalho

    # Título
    ws.merge_cells("A1:G1")
    titulo = ws.cell(row=1, column=1, value="Histórico de Análises de Crédito — Paulo Bio Imóveis")
    titulo.fill = _fill(_ORANGE)
    titulo.font = _font(bold=True, color=_WHITE, size=12)
    titulo.alignment = _center()
    ws.row_dimensions[1].height = 24

    # Cabeçalho
    for col_idx, (label, width) in enumerate(_COLUNAS_HISTORICO, start=1):
        _header_style(ws, row=2, col=col_idx, value=label, width=width)
    ws.row_dimensions[2].height = 20

    # Dados
    for row_offset, reg in enumerate(registros, start=3):
        aluguel_raw = reg.get("aluguel") or 0
        try:
            aluguel = float(aluguel_raw)
        except (TypeError, ValueError):
            aluguel = 0.0

        status = str(reg.get("status", "—"))
        cor = _cor_status(status)

        valores = [
            _fmt_data(reg.get("created_at")),
            reg.get("empresa") or "—",
            reg.get("cnpj") or "—",
            reg.get("imovel") or "—",
            _fmt_brl(aluguel),
            status,
            reg.get("usuario_nome") or "—",
        ]

        for col_idx, val in enumerate(valores, start=1):
            align = "center" if col_idx in (1, 3, 5, 6) else "left"
            color = cor if col_idx == 6 else None
            bold_status = col_idx == 6
            _data_cell(ws, row=row_offset, col=col_idx, value=val, align=align, color=color)
            if bold_status:
                ws.cell(row=row_offset, column=col_idx).font = _font(bold=True, color=cor)

        ws.row_dimensions[row_offset].height = 18


# ── Aba Resumo ────────────────────────────────────────────────────────────────

def _construir_aba_resumo(ws, registros: list[dict]) -> None:
    ws.title = "Resumo"

    def _status_key(status: str) -> str:
        s = str(status).upper()
        if "REPROVADO" in s:
            return "reprovado"
        if "RESSALVA" in s:
            return "ressalva"
        if "APROVADO" in s:
            return "aprovado"
        return "outro"

    # Agrupa por analista
    por_analista: dict[str, dict] = {}
    for reg in registros:
        analista = reg.get("usuario_nome") or "Desconhecido"
        if analista not in por_analista:
            por_analista[analista] = {"aprovado": 0, "reprovado": 0, "ressalva": 0, "outro": 0, "vgv": 0.0}
        key = _status_key(reg.get("status", ""))
        por_analista[analista][key] += 1
        try:
            por_analista[analista]["vgv"] += float(reg.get("aluguel") or 0)
        except (TypeError, ValueError):
            pass

    # Título
    ws.merge_cells("A1:G1")
    titulo = ws.cell(row=1, column=1, value="Resumo por Analista — Paulo Bio Imóveis")
    titulo.fill = _fill(_ORANGE)
    titulo.font = _font(bold=True, color=_WHITE, size=12)
    titulo.alignment = _center()
    ws.row_dimensions[1].height = 24

    cabecalhos = [
        ("Analista",       22),
        ("Total",          10),
        ("Aprovados",      12),
        ("Ressalvas",      12),
        ("Reprovados",     12),
        ("Taxa Aprovação", 16),
        ("VGV Total",      18),
    ]
    for col_idx, (label, width) in enumerate(cabecalhos, start=1):
        _header_style(ws, row=2, col=col_idx, value=label, width=width)
    ws.row_dimensions[2].height = 20

    total_geral = {"aprovado": 0, "reprovado": 0, "ressalva": 0, "outro": 0, "vgv": 0.0}

    for row_offset, (analista, dados) in enumerate(sorted(por_analista.items()), start=3):
        total = sum(dados[k] for k in ("aprovado", "reprovado", "ressalva", "outro"))
        taxa = f"{(dados['aprovado'] / total * 100):.1f}%" if total > 0 else "—"

        for k in ("aprovado", "reprovado", "ressalva", "outro"):
            total_geral[k] += dados[k]
        total_geral["vgv"] += dados["vgv"]

        row_vals = [
            analista,
            total,
            dados["aprovado"],
            dados["ressalva"],
            dados["reprovado"],
            taxa,
            _fmt_brl(dados["vgv"]),
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            align = "center" if col_idx > 1 else "left"
            _data_cell(ws, row=row_offset, col=col_idx, value=val, align=align)
        ws.row_dimensions[row_offset].height = 18

    # Linha de totais
    total_row = len(por_analista) + 3
    total_total = sum(total_geral[k] for k in ("aprovado", "reprovado", "ressalva", "outro"))
    taxa_geral = f"{(total_geral['aprovado'] / total_total * 100):.1f}%" if total_total > 0 else "—"

    totais = ["TOTAL GERAL", total_total, total_geral["aprovado"], total_geral["ressalva"],
              total_geral["reprovado"], taxa_geral, _fmt_brl(total_geral["vgv"])]
    for col_idx, val in enumerate(totais, start=1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.fill = _fill(_NAVY)
        cell.font = _font(bold=True, color=_WHITE)
        cell.alignment = _center() if col_idx > 1 else _left()
        cell.border = _border()
    ws.row_dimensions[total_row].height = 20


# ── Função pública ────────────────────────────────────────────────────────────

def gerar_excel_bytes(registros: list[dict]) -> bytes:
    """
    Recebe a lista de registros do Supabase (já filtrados) e retorna
    os bytes de um arquivo .xlsx pronto para download.
    """
    wb = openpyxl.Workbook()

    # Remove a aba padrão criada pelo openpyxl
    wb.remove(wb.active)  # type: ignore[arg-type]

    ws_historico = wb.create_sheet("Histórico")
    _construir_aba_historico(ws_historico, registros)

    ws_resumo = wb.create_sheet("Resumo")
    _construir_aba_resumo(ws_resumo, registros)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
